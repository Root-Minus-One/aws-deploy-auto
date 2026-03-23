"""File processing service for PDF, Word, Excel, and images."""

import io
import os
import uuid
import logging
import mimetypes
from pathlib import Path
from config import get_settings

logger = logging.getLogger(__name__)
settings = get_settings()

# Supported MIME types
SUPPORTED_TYPES = {
    # Images
    "image/jpeg",
    "image/png",
    "image/gif",
    "image/webp",
    # Documents
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",  # .docx
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/msword",  # .doc
    "application/vnd.ms-excel",  # .xls
    # Text
    "text/plain",
    "text/csv",
}


def get_upload_dir() -> Path:
    """Create and return the upload directory."""
    upload_dir = Path(settings.UPLOAD_DIR)
    upload_dir.mkdir(parents=True, exist_ok=True)
    return upload_dir


def validate_file(filename: str, content_type: str, size: int) -> None:
    """Validate uploaded file type and size."""
    if content_type not in SUPPORTED_TYPES:
        # Try to guess from filename
        guessed_type, _ = mimetypes.guess_type(filename)
        if guessed_type not in SUPPORTED_TYPES:
            raise ValueError(
                f"Unsupported file type: {content_type}. "
                f"Supported types: images, PDF, Word (.docx), Excel (.xlsx), text, CSV"
            )

    if size > settings.MAX_FILE_SIZE:
        max_mb = settings.MAX_FILE_SIZE / (1024 * 1024)
        raise ValueError(f"File too large. Maximum size is {max_mb:.0f}MB")


async def save_file(
    file_bytes: bytes,
    filename: str,
    content_type: str,
) -> dict:
    """Save file to disk and return metadata."""
    file_id = str(uuid.uuid4())
    ext = Path(filename).suffix
    stored_name = f"{file_id}{ext}"
    upload_dir = get_upload_dir()
    file_path = upload_dir / stored_name

    with open(file_path, "wb") as f:
        f.write(file_bytes)

    logger.info(f"File saved: {stored_name} ({len(file_bytes)} bytes)")

    return {
        "file_id": file_id,
        "filename": filename,
        "stored_name": stored_name,
        "content_type": content_type,
        "size": len(file_bytes),
        "path": str(file_path),
    }


async def get_file(file_id: str, db) -> dict | None:
    """Get file metadata from database."""
    file_doc = await db.files.find_one({"file_id": file_id})
    return file_doc


async def get_file_bytes(stored_name: str) -> bytes | None:
    """Read file bytes from disk."""
    upload_dir = get_upload_dir()
    file_path = upload_dir / stored_name
    if not file_path.exists():
        return None
    with open(file_path, "rb") as f:
        return f.read()


def extract_text_from_file(file_bytes: bytes, content_type: str) -> str:
    """Extract text content from supported file types for context."""
    try:
        if content_type == "text/plain" or content_type == "text/csv":
            return file_bytes.decode("utf-8", errors="replace")

        elif content_type == "application/pdf":
            import PyPDF2
            reader = PyPDF2.PdfReader(io.BytesIO(file_bytes))
            text = ""
            for page in reader.pages:
                text += page.extract_text() or ""
            return text

        elif content_type in (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/msword",
        ):
            import docx
            doc = docx.Document(io.BytesIO(file_bytes))
            return "\n".join(p.text for p in doc.paragraphs)

        elif content_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            text_parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text_parts.append(f"Sheet: {sheet}")
                for row in ws.iter_rows(values_only=True):
                    row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                    text_parts.append(row_text)
            return "\n".join(text_parts)

        else:
            return ""

    except Exception as e:
        logger.error(f"Text extraction error for {content_type}: {e}")
        return ""
